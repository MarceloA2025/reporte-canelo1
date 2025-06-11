from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# Ruta al archivo
archivo = r"C:\One Drive Hotmail\OneDrive\Documentos\Python VSCode\REPORTE WEB\HEC mensuales 2025.xlsx"

# Abrir archivo
wb = load_workbook(archivo)
ws = wb["Pluviometria"]

# Detectar hasta dónde llegan los datos reales desde la fila 128 en adelante
fila_inicio = 128
fila_actual = fila_inicio

while ws[f"C{fila_actual}"].value and ws[f"D{fila_actual}"].value:
    fila_actual += 1

fila_final = fila_actual - 1  # Última fila con datos

# Definir rango de tabla
celda_inicio = f"C{fila_inicio}"
celda_fin = f"D{fila_final}"
rango = f"{celda_inicio}:{celda_fin}"

# Crear tabla
tabla = Table(displayName="Precipitaciones", ref=rango)

# Estilo visual
estilo = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
tabla.tableStyleInfo = estilo

# Agregar tabla
ws.add_table(tabla)

# Guardar
wb.save(archivo)

print(f"✅ Tabla 'Precipitaciones' creada desde {celda_inicio} hasta {celda_fin}.")
