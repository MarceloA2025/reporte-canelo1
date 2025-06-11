from openpyxl import load_workbook

archivo = "HEC mensuales 2025.xlsx"  # Asegúrate de que esté en la misma carpeta

wb = load_workbook(archivo, data_only=True)

print("🔍 Revisando tablas en el archivo...\n")
for hoja in wb.sheetnames:
    ws = wb[hoja]
    tablas = ws._tables
    if tablas:
        print(f"📄 Hoja '{hoja}' contiene las siguientes tablas:")
        for t in tablas:
            if hasattr(t, "name") and hasattr(t, "ref"):
                print(f"   🔹 {t.name} (rango: {t.ref})")
            else:
                print(f"   🔹 Tabla mal definida (probablemente solo el nombre): {t}")
    else:
        print(f"📄 Hoja '{hoja}' no contiene tablas.")
